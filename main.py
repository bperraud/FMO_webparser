import requests
from bs4 import BeautifulSoup
import openpyxl
import os

def create_db_from_url(URL : str, nb_pages : int, filename : str) :
	filename += ".xlsx"
	if os.path.exists(filename):
		workbook = openpyxl.load_workbook(filename)
	else:
		workbook = openpyxl.Workbook()

	sheet = workbook.active
	header = ["Name", "Country", "Currency", "Amount", "Sector", "Date"]
	header_index = {column: index for index, column in enumerate(header)}
	sheet.delete_rows(1, sheet.max_row)
	sheet.append(header)

	page = 1
	print("parsing...")

	while (page <= nb_pages) :
		response = requests.get(URL + "&page=" + str(page))
		if response.status_code == 200:
			soup = BeautifulSoup(response.text, 'html.parser')
			span_elements = soup.find_all('li', class_='ProjectList__item')

			rows = 1

			for parent_span in span_elements:
				title_element = parent_span.find('h3', class_='ProjectList__projectTitle')
				data_row = [""] * len(header_index)

				if title_element:
					data_row[header_index["Name"]] = title_element.text

				child_span = parent_span.find_all('span', class_='ProjectList__projectExtras')

				for span in child_span:
					financing = span.find('span', class_='fmo-financing')
					if financing:
						parts = financing.text.split(' ')
						currency = parts[0].strip()
						amount = parts[1].strip()
						data_row[header_index["Currency"]] = currency
						data_row[header_index["Amount"]] = amount

					lines = span.text.splitlines()
					for line in lines:
						if ':' in line:
							parts = line.split(':')
							title = parts[0].strip()
							value = parts[1].strip()
							if (title in header_index) :
								data_row[header_index[title]] = value
					sheet.append(data_row)
					rows += 1
		else:
			print(f"Échec de la requête. Code d'état : {response.status_code}")

		page += 1

	# Save the Excel file to a specific path
	workbook.save(filename)
	print("finished for", filename)

create_db_from_url("https://www.fmo.nl/world-map?search=&region=&year=&fund%5B%5D=4", 6, "building_prospects")
create_db_from_url("https://www.fmo.nl/world-map?search=&region=&year=&fund%5B%5D=1", 4, "access_to_energy_fund")
